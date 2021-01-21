import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '회원정보'
header_titles = ['이름', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1, value=title)

members = [('kei', '010-1234-1234'), ('hong','010-4321-4321')]
row_num = 2
for r, member in enumerate(members):
    for c, v in enumerate(member):
        sheet.cell(row=row_num+r, column=c+1, value=v)

wb.save('./members.xlsx')
wb.close()